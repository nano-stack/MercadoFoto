#!/usr/bin/env python3
"""
email_publisher.py
==================
Lee correos en la casilla configurada, detecta adjuntos Excel (.xlsx / .xls)
y publica automáticamente cada fila como un producto en OkVenta/MercadoFoto.

Uso:
    python email_publisher.py                  # bucle continuo
    python email_publisher.py --once           # procesa una sola vez y sale
    python email_publisher.py --config ruta.json

Configuración: config_email.json (mismo directorio)
"""

import argparse
import imaplib
import email
import io
import json
import os
import sys
import time
import traceback
from email.header import decode_header

from services.email_service import enviar_confirmacion_carga_masiva
from database.users import obtener_ubicacion_usuario, obtener_usuario_por_email
from pathlib import Path
from typing import Optional

import openpyxl
import requests

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
DEFAULT_CONFIG = BASE_DIR / "config_email.json"


def cargar_config(path: Path) -> dict:
    # En Render no existe config_email.json — se usan variables de entorno
    if not path.exists():
        imap_email    = os.environ.get("IMAP_EMAIL")
        imap_password = os.environ.get("IMAP_PASSWORD")
        api_url       = os.environ.get("API_URL")
        if not imap_email or not imap_password or not api_url:
            raise EnvironmentError(
                "Faltan variables de entorno: IMAP_EMAIL, IMAP_PASSWORD, API_URL"
            )
        return {
            "imap_server":            os.environ.get("IMAP_SERVER", "mail.galmar.cl"),
            "imap_port":              int(os.environ.get("IMAP_PORT", "993")),
            "email":                  imap_email,
            "password":               imap_password,
            "api_url":                api_url,
            "check_interval_seconds": int(os.environ.get("CHECK_INTERVAL", "300")),
            "marca_leido":            True,
            "carpeta":                os.environ.get("IMAP_CARPETA", "INBOX"),
        }

    with open(path, encoding="utf-8") as f:
        cfg = json.load(f)
    required = ["imap_server", "imap_port", "email", "password", "api_url"]
    for key in required:
        if key not in cfg:
            raise ValueError(f"Falta '{key}' en {path}")
    cfg.setdefault("check_interval_seconds", 300)
    cfg.setdefault("marca_leido", True)
    cfg.setdefault("carpeta", "INBOX")
    return cfg


# ─────────────────────────────────────────────
# DETECCIÓN FLEXIBLE DE COLUMNAS
# ─────────────────────────────────────────────

# Para cada campo destino, lista de posibles nombres de cabecera (en minúsculas)
COLUMN_VARIANTS: dict[str, list[str]] = {
    "titulo":           ["titulo", "título", "nombre", "producto", "name", "title"],
    "precio":           ["precio", "price", "valor", "costo", "monto"],
    "estado":           ["estado", "condicion", "condición", "state", "condition"],
    "descripcion":      ["descripcion", "descripción", "description", "detalle", "desc", "observaciones"],
    "categoria":        ["categoria", "categoría", "category", "tipo", "rubro"],
    "subcategoria":     ["subcategoria", "subcategoría", "subcategory", "subtipo", "subrubro"],
    "imagen":           ["imagen_url", "imagen", "image", "foto", "photo", "imagen1", "foto1", "url_imagen", "link_imagen"],
    "imagen2":          ["imagen_url2", "imagen2", "foto2", "image2", "photo2"],
    "imagen3":          ["imagen_url3", "imagen3", "foto3", "image3", "photo3"],
    "sku":              ["sku", "codigo_interno", "cod_interno", "referencia"],
    "codigo_universal": ["codigo universal", "codigo_universal", "barcode", "ean", "upc", "gtin"],
}


def _normalizar(texto: str) -> str:
    """Minúsculas, sin espacios extra, sin tildes problemáticas."""
    return str(texto).strip().lower().replace("  ", " ")


def detectar_mapa_columnas(cabeceras: list[str]) -> dict[str, int]:
    """
    Dado el listado de cabeceras de la hoja, devuelve un dict
    {campo_destino: índice_columna}.
    """
    mapa: dict[str, int] = {}
    for idx, cab in enumerate(cabeceras):
        cab_norm = _normalizar(cab)
        for campo, variantes in COLUMN_VARIANTS.items():
            if campo not in mapa and cab_norm in variantes:
                mapa[campo] = idx
                break
    return mapa


def buscar_fila_cabecera(ws) -> Optional[int]:
    """
    Recorre las primeras 10 filas buscando aquella que contenga al menos
    'titulo'/'título'/'nombre' Y 'precio'. Devuelve el número de fila (1-based).
    """
    imprescindibles = set(COLUMN_VARIANTS["titulo"]) | set(COLUMN_VARIANTS["precio"])
    for row_num in range(1, 11):
        valores = [_normalizar(str(ws.cell(row=row_num, column=c).value or ""))
                   for c in range(1, ws.max_column + 1)]
        if any(v in imprescindibles for v in valores):
            return row_num
    return None


# ─────────────────────────────────────────────
# LECTURA DEL EXCEL
# ─────────────────────────────────────────────

def leer_productos_excel(
    contenido_excel: bytes,
    imagenes_adjuntas: dict[str, bytes],
    email_remitente: str = "",
) -> list[dict]:
    """
    Parsea el Excel y devuelve una lista de dicts con los campos del producto.
    `imagenes_adjuntas` es {nombre_archivo_lower: bytes} de las imágenes del correo.
    """
    wb = openpyxl.load_workbook(io.BytesIO(contenido_excel), data_only=True)
    ws = wb.active

    fila_cab = buscar_fila_cabecera(ws)
    if fila_cab is None:
        print("  ⚠  No se detectó fila de cabeceras en el Excel.")
        return []

    # Leer cabeceras
    cabeceras = [
        str(ws.cell(row=fila_cab, column=c).value or "")
        for c in range(1, ws.max_column + 1)
    ]
    mapa = detectar_mapa_columnas(cabeceras)
    print(f"  Mapa de columnas detectado: {mapa}")

    if "titulo" not in mapa or "precio" not in mapa:
        print("  ⚠  No se encontraron columnas obligatorias (titulo, precio).")
        return []

    productos = []
    for row_num in range(fila_cab + 1, ws.max_row + 1):
        fila = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]

        # Ignorar filas vacías
        if all(v is None or str(v).strip() == "" for v in fila):
            continue

        def get(campo: str, default=None):
            if campo in mapa:
                val = fila[mapa[campo]]
                return val if val is not None else default
            return default

        titulo = str(get("titulo", "")).strip()
        if not titulo:
            continue

        try:
            precio = float(str(get("precio", 0)).replace(",", ".").replace("$", "").strip())
        except (ValueError, TypeError):
            print(f"  ⚠  Fila {row_num}: precio inválido, se omite.")
            continue

        descripcion  = str(get("descripcion", titulo)).strip() or titulo
        categoria    = str(get("categoria",  "")).strip() or None
        subcategoria = str(get("subcategoria","")).strip() or None

        # Usuario identificado por email del remitente
        user_id = None
        if email_remitente:
            usuario = obtener_usuario_por_email(email_remitente)
            if usuario:
                user_id = usuario["id"]

        if not user_id:
            print(f"  ⚠  Fila {row_num}: no se encontró usuario para '{email_remitente}', se omite.")
            continue

        # lat/lng desde la dirección del perfil del usuario
        lat = None
        lng = None
        ubicacion = obtener_ubicacion_usuario(user_id)
        if ubicacion:
            lat = ubicacion.get("lat")
            lng = ubicacion.get("lng")

        # Imágenes: leer URLs de la planilla
        imagen_urls: list[str] = []
        for campo_img in ["imagen", "imagen2", "imagen3"]:
            url = str(get(campo_img, "")).strip()
            if url and url.startswith("http"):
                imagen_urls.append(url)

        productos.append({
            "titulo":       titulo,
            "descripcion":  descripcion,
            "precio":       precio,
            "categoria":    categoria,
            "subcategoria": subcategoria,
            "user_id":      user_id,
            "lat":          lat,
            "lng":          lng,
            "imagen_urls":  imagen_urls,  # lista de URLs
        })

    return productos


# ─────────────────────────────────────────────
# PUBLICAR VÍA API
# ─────────────────────────────────────────────

def _descargar_imagen(url: str) -> Optional[bytes]:
    """Descarga una imagen desde una URL y retorna sus bytes."""
    try:
        resp = requests.get(url, timeout=30)
        if resp.status_code == 200 and resp.content:
            return resp.content
    except Exception as e:
        print(f"  ⚠  No se pudo descargar imagen {url}: {e}")
    return None


def publicar_producto(api_url: str, producto: dict) -> bool:
    """
    Descarga imágenes desde URLs y llama a POST /publicar.
    Devuelve True si tuvo éxito.
    """
    if not producto["user_id"]:
        print(f"  ⚠  '{producto['titulo']}': sin user_id, se omite.")
        return False

    # Descargar imágenes desde URLs
    imagen_urls = producto.get("imagen_urls", [])
    imagenes_bytes = []
    for url in imagen_urls[:4]:
        img = _descargar_imagen(url)
        if img:
            imagenes_bytes.append(img)

    if not imagenes_bytes:
        print(f"  ⚠  '{producto['titulo']}': sin imágenes válidas, se omite.")
        return False

    campos = {
        "titulo":      str(producto["titulo"]),
        "descripcion": str(producto["descripcion"]),
        "precio":      str(producto["precio"]),
        "user_id":     str(producto["user_id"]),
    }
    if producto.get("categoria"):
        campos["categoria"] = producto["categoria"]
    if producto.get("subcategoria"):
        campos["subcategoria"] = producto["subcategoria"]
    if producto.get("lat") is not None:
        campos["lat"] = str(producto["lat"])
    if producto.get("lng") is not None:
        campos["lng"] = str(producto["lng"])

    slots = ["file", "file2", "file3", "file4"]
    archivos = {}
    for i, img_bytes in enumerate(imagenes_bytes):
        archivos[slots[i]] = (f"imagen{i+1}.jpg", img_bytes, "image/jpeg")

    try:
        resp = requests.post(
            f"{api_url}/publicar",
            data=campos,
            files=archivos,
            timeout=120,
        )
        if resp.status_code == 200:
            data = resp.json()
            print(f"  ✓  Publicado: '{producto['titulo']}' → {data.get('imagen_url','')}")
            return True
        else:
            print(f"  ✗  Error {resp.status_code}: {resp.text[:200]}")
            return False
    except Exception as e:
        print(f"  ✗  Excepción al publicar '{producto['titulo']}': {e}")
        return False


# ─────────────────────────────────────────────
# LECTURA DE CORREOS
# ─────────────────────────────────────────────

def obtener_nombre_adjunto(part) -> str:
    """Decodifica el nombre del adjunto de forma segura."""
    filename = part.get_filename() or ""
    if filename:
        decoded_parts = decode_header(filename)
        partes = []
        for data, charset in decoded_parts:
            if isinstance(data, bytes):
                partes.append(data.decode(charset or "utf-8", errors="replace"))
            else:
                partes.append(data)
        filename = "".join(partes)
    return filename


def procesar_correo(msg, api_url: str, email_remitente: str = "") -> int:
    """
    Extrae adjuntos de un correo y publica los productos encontrados.
    Devuelve el número de productos publicados.
    """
    # Separar adjuntos Excel e imágenes
    excels: list[tuple[str, bytes]] = []          # (nombre, bytes)
    imagenes: dict[str, bytes] = {}               # {nombre_lower: bytes}

    for part in msg.walk():
        disposition = part.get_content_disposition() or ""
        content_type = part.get_content_type()
        filename = obtener_nombre_adjunto(part)
        filename_lower = filename.lower()

        payload = part.get_payload(decode=True)
        if payload is None:
            continue

        # Excel
        if (filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls") or
                "spreadsheet" in content_type or "excel" in content_type):
            excels.append((filename, payload))

        # Imagen
        elif (content_type.startswith("image/") or
              any(filename_lower.endswith(ext) for ext in [".jpg", ".jpeg", ".png", ".webp", ".gif"])):
            imagenes[filename_lower] = payload

    if not excels:
        return 0

    total_publicados = 0
    for nombre_excel, contenido_excel in excels:
        print(f"  Procesando Excel: {nombre_excel}")
        try:
            productos = leer_productos_excel(contenido_excel, imagenes, email_remitente)
            print(f"  → {len(productos)} producto(s) encontrado(s)")
            for p in productos:
                if publicar_producto(api_url, p):
                    total_publicados += 1
        except Exception as e:
            print(f"  ✗  Error procesando {nombre_excel}: {e}")
            traceback.print_exc()

    return total_publicados


def revisar_correos(cfg: dict) -> int:
    """
    Conecta al IMAP, procesa correos no leídos con adjuntos Excel.
    Devuelve el total de productos publicados.
    """
    server   = cfg["imap_server"]
    port     = cfg["imap_port"]
    usuario  = cfg["email"]
    password = cfg["password"]
    carpeta  = cfg["carpeta"]
    api_url  = cfg["api_url"].rstrip("/")
    marcar   = cfg["marca_leido"]

    print(f"\n[{time.strftime('%Y-%m-%d %H:%M:%S')}] Revisando {usuario}…")

    try:
        mail = imaplib.IMAP4_SSL(server, port)
        mail.login(usuario, password)
        mail.select(carpeta)

        # Buscar no leídos con adjunto (IMAP no filtra por adjunto nativamente,
        # pero filtramos tras leer el mensaje)
        status, data = mail.search(None, "UNSEEN")
        if status != "OK":
            print("  No se pudo buscar mensajes.")
            mail.logout()
            return 0

        ids = data[0].split()
        if not ids:
            print("  Sin mensajes nuevos.")
            mail.logout()
            return 0

        print(f"  {len(ids)} mensaje(s) sin leer.")
        total = 0

        for uid in ids:
            res, msg_data = mail.fetch(uid, "(RFC822)")
            if res != "OK":
                continue

            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)

            asunto_raw = msg.get("Subject", "")
            decoded = decode_header(asunto_raw)
            asunto = ""
            for data, charset in decoded:
                if isinstance(data, bytes):
                    asunto += data.decode(charset or "utf-8", errors="replace")
                else:
                    asunto += data
            print(f"\n  ✉  Asunto: {asunto}")

            # Obtener remitente para enviar confirmación
            remitente_raw = msg.get("From", "")
            # Extraer solo el email del campo From ("Nombre <email@x.com>" → "email@x.com")
            if "<" in remitente_raw and ">" in remitente_raw:
                email_remitente = remitente_raw.split("<")[1].split(">")[0].strip()
            else:
                email_remitente = remitente_raw.strip()

            publicados = procesar_correo(msg, api_url, email_remitente)
            total += publicados

            if marcar and publicados > 0:
                mail.store(uid, "+FLAGS", "\\Seen")
                print(f"  → Marcado como leído ({publicados} publicado(s)).")
                # Enviar confirmación al remitente
                if email_remitente:
                    username = email_remitente.split("@")[0]
                    enviar_confirmacion_carga_masiva(
                        email_destino=email_remitente,
                        username=username,
                        total_publicados=publicados,
                    )
            elif publicados == 0:
                print("  → Sin productos publicados (sin Excel válido o sin datos).")

        mail.logout()
        return total

    except imaplib.IMAP4.error as e:
        print(f"  ✗  Error IMAP: {e}")
        return 0
    except Exception as e:
        print(f"  ✗  Error inesperado: {e}")
        traceback.print_exc()
        return 0


# ─────────────────────────────────────────────
# PUNTO DE ENTRADA
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Email auto-publisher para OkVenta")
    parser.add_argument("--once",   action="store_true", help="Procesar una vez y salir")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="Ruta al JSON de configuración")
    args = parser.parse_args()

    config_path = Path(args.config)
    if not config_path.exists():
        print(f"✗ No se encontró la configuración: {config_path}")
        print("  Crea config_email.json con: imap_server, imap_port, email, password, api_url")
        sys.exit(1)

    cfg = cargar_config(config_path)
    intervalo = cfg["check_interval_seconds"]

    print("=" * 55)
    print("  OkVenta — Email Auto-Publisher")
    print(f"  Cuenta  : {cfg['email']}")
    print(f"  Servidor: {cfg['imap_server']}:{cfg['imap_port']}")
    print(f"  API     : {cfg['api_url']}")
    print(f"  Intervalo: {intervalo}s" if not args.once else "  Modo: una sola ejecución")
    print("=" * 55)

    if args.once:
        revisar_correos(cfg)
        return

    # Bucle continuo
    while True:
        try:
            revisar_correos(cfg)
        except KeyboardInterrupt:
            print("\nDetenido por el usuario.")
            break
        except Exception as e:
            print(f"Error en bucle principal: {e}")
            traceback.print_exc()

        print(f"  Próxima revisión en {intervalo}s… (Ctrl+C para salir)\n")
        try:
            time.sleep(intervalo)
        except KeyboardInterrupt:
            print("\nDetenido por el usuario.")
            break


if __name__ == "__main__":
    main()
